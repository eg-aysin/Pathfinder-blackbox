import os
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
import httpx
from pydantic import BaseModel
from pathlib import Path
from openpyxl import load_workbook, Workbook
from io import BytesIO

app = FastAPI(title="EnerKey")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

PATHFINDER_URL = os.getenv("PATHFINDER_URL", "http://localhost:8001")
HTML_FILE = Path(__file__).parent / "templates" / "index.html"


class ImportRequest(BaseModel):
    account_names: list[str]


@app.get("/", response_class=HTMLResponse)
async def index():
    return HTMLResponse(content=HTML_FILE.read_text(encoding="utf-8"))


@app.post("/api/import")
async def mass_import(data: ImportRequest):
    results = []
    async with httpx.AsyncClient(timeout=120.0) as client:
        for name in data.account_names:
            response = await client.post(
                f"{PATHFINDER_URL}/api/import",
                json={"facility_name": name},
            )
            if response.status_code != 200:
                return JSONResponse(
                    status_code=502,
                    content={"error": f"Pathfinder returned {response.status_code}: {response.text}"},
                )
            results.append((name, response.content))

    # Merge all results into one sheet
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "Merged Accounts"

    for idx, (name, content) in enumerate(results):
        wb = load_workbook(filename=BytesIO(content))
        ws = wb.active
        # Add account name as separator row
        merged_ws.append([f"Account: {name}"])
        for row in ws.iter_rows(values_only=True):
            merged_ws.append(row)
        # Add empty row between accounts
        merged_ws.append([])

    output_stream = BytesIO()
    merged_wb.save(output_stream)
    output_stream.seek(0)

    output_filename = "output_multi.xlsx" if len(data.account_names) > 1 else f'output_{data.account_names[0].replace(" ", "_").replace("/", "-")}.xlsx'
    return StreamingResponse(
        iter([output_stream.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{output_filename}"'
        }
    )


@app.get("/health")
def health():
    return {"status": "ok", "service": "enerkey"}
